import json
import os
from collections import Counter
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Union

from openpyxl import load_workbook

from config.constants import SALES_SHEET_NAME
from repositories.sales_repository import RepositorioVendasSqlModel
from utils.parsers import parse_date, parse_float

class ExtrairDadosProdutos:
    """Responsável por extrair dados de produtos a partir de uma planilha."""

    def __init__(self, workbook, sheet_name: str):
        self.workbook = workbook
        self.sheet_name = sheet_name

    def executar(self):
        worksheet = self.workbook[self.sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        dados = []

        if not rows:
            return dados

        headers = list(rows[0])
        id_idx = headers.index('ID Venda')
        data_venda_idx = headers.index('Data da Venda')
        produto_idx = headers.index('Produto')
        categoria_idx = headers.index('Categoria')
        preco_unitario_idx = headers.index('Preço Unitário')
        quantidade_vendida_idx = headers.index('Quantidade Vendida')
        estoque_atual_idx = headers.index('Estoque Atual')

        for row in rows[1:]:
            id_venda = row[id_idx]
            data_venda = parse_date(row[data_venda_idx])
            produto = row[produto_idx]
            categoria = row[categoria_idx]
            preco_unitario = parse_float(row[preco_unitario_idx])
            quantidade_vendida = parse_float(row[quantidade_vendida_idx])
            faturamento_total = round(preco_unitario * quantidade_vendida, 2)
            estoque_atual = row[estoque_atual_idx]

            if id_venda is None or produto is None:
                continue

            dados.append((
                id_venda,
                data_venda,
                produto,
                categoria,
                preco_unitario,
                quantidade_vendida,
                faturamento_total,
                estoque_atual,
            ))

        return dados


class ResumoVendas:
    """Responsável por resumir vendas e identificar tendências."""

    def __init__(self, workbook, sheet_name: str):
        self.workbook = workbook
        self.sheet_name = sheet_name

    def extrair_tendencias(self, data_inicial=None, data_final=None):
        extrator = ExtrairDadosProdutos(self.workbook, self.sheet_name)
        dados_produtos = extrator.executar()

        if data_inicial:
            data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d')
            dados_produtos = [
                row for row in dados_produtos
                if datetime.strptime(row[1], '%Y-%m-%d') >= data_inicial
            ]

        if data_final:
            data_final = datetime.strptime(data_final, '%Y-%m-%d')
            dados_produtos = [
                row for row in dados_produtos
                if datetime.strptime(row[1], '%Y-%m-%d') <= data_final
            ]

        totais_produtos = Counter()
        totais_categorias = Counter()

        for row in dados_produtos:
            quantidade_vendida = float(row[5] or 0)
            produto = row[2]
            categoria = row[3]

            totais_produtos[produto] += quantidade_vendida
            totais_categorias[categoria] += quantidade_vendida

        mais_vendidos = sorted(totais_produtos.items(), key=lambda item: (-item[1], item[0]))[:1]
        menos_vendidos = sorted(totais_produtos.items(), key=lambda item: (item[1], item[0]))[:1]
        categoria_mais_vendida = sorted(totais_categorias.items(), key=lambda item: (-item[1], item[0]))[:1]

        return {
            'mais_vendidos': [{'produto': produto, 'quantidade': quantidade} for produto, quantidade in mais_vendidos],
            'menos_vendidos': [{'produto': produto, 'quantidade': quantidade} for produto, quantidade in menos_vendidos],
            'categoria_mais_vendida': {'categoria': categoria_mais_vendida[0][0], 'quantidade': categoria_mais_vendida[0][1]} if categoria_mais_vendida else {}
        }


def _carregar_workbook(workbook_or_path: Union[object, str], sheet_name: Optional[str] = None):
    if isinstance(workbook_or_path, str):
        if not os.path.exists(workbook_or_path):
            raise FileNotFoundError(f'Arquivo não encontrado: {workbook_or_path}')
        workbook = load_workbook(filename=workbook_or_path, data_only=True)
        return workbook, sheet_name or SALES_SHEET_NAME

    if hasattr(workbook_or_path, 'sheetnames'):
        return workbook_or_path, sheet_name or workbook_or_path.sheetnames[0]

    raise TypeError('Esperava um workbook do openpyxl ou um caminho para um arquivo .xlsx')


def _obter_repositorio(database_url: Optional[str] = None) -> RepositorioVendasSqlModel:
    return RepositorioVendasSqlModel(database_url=database_url)


def extract_produtos(workbook_or_path: Union[object, str], sheet_name: Optional[str] = None):
    workbook, resolved_sheet_name = _carregar_workbook(workbook_or_path, sheet_name)
    return ExtrairDadosProdutos(workbook, resolved_sheet_name).executar()


def resumir_vendas(workbook_or_path: Union[object, str], sheet_name: Optional[str] = None, data_inicial=None, data_final=None):
    workbook, resolved_sheet_name = _carregar_workbook(workbook_or_path, sheet_name)
    return ResumoVendas(workbook, resolved_sheet_name).extrair_tendencias(data_inicial, data_final)


def persistir_dados(workbook_or_path: Union[object, str], sheet_name: Optional[str] = None, database_url: Optional[str] = None):
    dados = extract_produtos(workbook_or_path, sheet_name)
    repositorio = _obter_repositorio(database_url)
    return repositorio.persistir(dados, database_url=database_url)


def consultar_vendas(database_url: Optional[str] = None, limit: int = 100) -> List[Dict[str, Any]]:
    repositorio = _obter_repositorio(database_url)
    return repositorio.consultar(database_url=database_url, limit=limit)


if __name__ == '__main__':
    workbook = load_workbook(filename='files/sales_worksheet.xlsx')
    sheet_name = SALES_SHEET_NAME

    resultado = extract_produtos(workbook, sheet_name)
    resumo = resumir_vendas(workbook, sheet_name)

    print('Resultado extraído:')
    print(json.dumps(resultado, indent=4, ensure_ascii=False))
    print('Resumo de vendas:')
    print(json.dumps(resumo, indent=4, ensure_ascii=False))